from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from app.services.database import get_all_complaints, get_branch_counts, get_total_complaints


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, row_num: int = 1):
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_excel_report(output_dir: str = "data") -> str:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Hisobot"
    summary_ws.append(["Ko'rsatkich", "Qiymat"])
    _style_header(summary_ws)
    summary_ws.append(["Umumiy shikoyatlar soni", get_total_complaints()])

    branch_ws = wb.create_sheet("Filiallar")
    branch_ws.append(["Filial", "Shikoyatlar soni"])
    _style_header(branch_ws)
    for row in get_branch_counts():
        branch_ws.append([row["branch"], row["total"]])

    detail_ws = wb.create_sheet("Batafsil")
    detail_ws.append([
        "ID",
        "Sana",
        "Filial",
        "Kim ustidan",
        "Shikoyat",
        "Foydalanuvchi",
        "Username",
        "Telegram ID",
        "Telefon raqam",
    ])
    _style_header(detail_ws)

    for row in get_all_complaints():
        detail_ws.append([
            row["id"],
            row["created_at"],
            row["branch"],
            row["target_role"],
            row["complaint_text"],
            row["full_name"],
            row["username"] or "-",
            row["user_id"],
            row["phone"] or "-",
        ])

    for ws in [summary_ws, branch_ws, detail_ws]:
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    filename = f"complaints_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = str(Path(output_dir) / filename)
    wb.save(file_path)
    return file_path
